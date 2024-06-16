import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

# Lista plików do przetworzenia
files = ['1_kolejka.xlsx', '2_kolejka.xlsx', '3_kolejka.xlsx', '1_8_playoff.xlsx', '1_4_playoff.xlsx',
         '1_2_playoff.xlsx', 'final.xlsx']

# Słownik do przechowywania sumy punktów dla każdego emaila oraz szczegółowych wyników
total_scores = {}
detailed_scores = {}

# Funkcja do przetwarzania pliku
def process_file(file_path, file_name):
    try:
        # Wczytaj dane z pliku Excel
        df_answers = pd.read_excel(file_path, sheet_name='Answers')
        df_correct = pd.read_excel(file_path, sheet_name='Correct_answers')

        # Zakładamy, że poprawne odpowiedzi są w jednym wierszu w arkuszu 'Correct_answers'
        correct_answers = df_correct.iloc[0].to_dict()

        # Przejdź przez odpowiedzi i porównaj z poprawnymi odpowiedziami
        for index, row in df_answers.iterrows():
            email = row['Adres e-mail']
            if email not in total_scores:
                total_scores[email] = 0
                detailed_scores[email] = []

            score = 0
            email_details = []

            for column in range(2, len(df_answers.columns), 2):  # Pomijamy kolumny 'Sygnatura czasowa' i 'Adres e-mail'
                team1_column = df_answers.columns[column]
                team2_column = df_answers.columns[column + 1]

                # Usuwamy drużynę w nawiasach z nazwy kolumny
                match = team1_column.split(' [')[0]

                user_team1_score = row[team1_column]
                user_team2_score = row[team2_column]

                correct_team1_score = correct_answers[team1_column]
                correct_team2_score = correct_answers[team2_column]

                # Sprawdź dokładny wynik
                if user_team1_score == correct_team1_score and user_team2_score == correct_team2_score:
                    score += 3
                    email_details.append((file_name.replace('.xlsx', '').replace('_', ' '), match, user_team1_score, user_team2_score, 3))
                # Sprawdź poprawne wytypowanie zwycięzcy lub remis
                elif (user_team1_score > user_team2_score and correct_team1_score > correct_team2_score) or \
                        (user_team1_score < user_team2_score and correct_team1_score < correct_team2_score) or \
                        (user_team1_score == user_team2_score and correct_team1_score == correct_team2_score):
                    score += 1
                    email_details.append((file_name.replace('.xlsx', '').replace('_', ' '), match, user_team1_score, user_team2_score, 1))
                else:
                    email_details.append((file_name.replace('.xlsx', '').replace('_', ' '), match, user_team1_score, user_team2_score, 0))

            total_scores[email] += score
            detailed_scores[email].extend(email_details)

    except Exception as e:
        print(f'Błąd przetwarzania pliku {file_path}: {e}')

if __name__ == "__main__":
    # Przetwarzaj każdy plik
    for file in files:
        if os.path.exists(file):
            process_file(file, os.path.basename(file))
        else:
            print(f'Plik {file} nie istnieje.')

    # Utwórz nową tabelę tylko z kolumnami 'Adres e-mail' i 'Punkty'
    df_results = pd.DataFrame(list(total_scores.items()), columns=['Adres e-mail', 'Punkty'])

    # Sortuj tabelę według kolumny 'Punkty' malejąco
    df_results = df_results.sort_values(by='Punkty', ascending=False)

    # Zapisz wyniki do nowego pliku Excel
    output_file_path = 'wyniki_z_punktami.xlsx'
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        # Zapisz ogólne wyniki
        df_results.to_excel(writer, sheet_name='Wyniki', index=False)

        # Zapisz szczegółowe wyniki dla każdego emaila
        for email, details in detailed_scores.items():
            df_details = pd.DataFrame(details,
                                      columns=['Kolejka/Tura', 'Mecz', 'Obstawiony', 'Wynik', 'Zdobyte punkty'])
            df_details.to_excel(writer, sheet_name=email, index=False)

        # Ustawienia kolorów dla punktów
        green_fill = PatternFill(start_color="33FF33", end_color="33FF33", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF57", end_color="FFFF57", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Otwórz zapisany plik, aby zmienić kolory komórek
        book = writer.book
        for email in detailed_scores.keys():
            sheet = book[email]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
                points = row[4].value
                if points == 3:
                    for cell in row:
                        cell.fill = green_fill
                elif points == 1:
                    for cell in row:
                        cell.fill = yellow_fill
                elif points == 0:
                    for cell in row:
                        cell.fill = white_fill

    print(f'Wyniki zostały zapisane w pliku: {output_file_path}')
